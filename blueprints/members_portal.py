# -*- coding: utf-8 -*-
"""
Cổng nội bộ Members - /members, /members/verify, /members/logout, /api/members
"""
import logging
import os
import re
import unicodedata
from datetime import date, datetime
from io import BytesIO
from pathlib import Path

from flask import Blueprint, redirect, render_template, request, jsonify, session, send_file

from extensions import rate_limit

logger = logging.getLogger(__name__)
members_portal_bp = Blueprint('members_portal', __name__)

# Bulk Update SLL: bật/tắt endpoint /api/members/bulk-update-sll (nút Update SLL trên members.html).
BULK_UPDATE_SLL_ENABLED = True


@members_portal_bp.route('/members', strict_slashes=False)
def members():
    """Trang danh sách thành viên - hiển thị cổng đăng nhập hoặc trang members."""
    if not session.get('members_gate_ok'):
        return render_template('members_gate.html')
    from services.members_service import get_members_password
    members_password = get_members_password()
    gate_username = session.get('members_gate_user', '')
    return render_template('members.html', members_password=members_password or '', gate_username=gate_username)


@members_portal_bp.route('/members/verify', methods=['POST'])
@rate_limit("25 per minute")
def members_verify():
    """API xác thực đăng nhập cho cổng Members. Nhận JSON hoặc form."""
    try:
        data = request.get_json(silent=True) if request.is_json else {}
        if not data:
            data = {
                'username': (request.form.get('username') or '').strip(),
                'password': (request.form.get('password') or '').strip(),
            }
        else:
            data = {'username': (data.get('username') or '').strip(), 'password': (data.get('password') or '').strip()}
        username = data.get('username', '')
        password = data.get('password', '')
        if not username or not password:
            return (jsonify({'success': False, 'error': 'Tên đăng nhập và mật khẩu không được để trống'}), 400)
        from app import validate_tbqc_gate
        if validate_tbqc_gate(username, password):
            session['members_gate_ok'] = True
            session['members_gate_user'] = username
            session.permanent = True
            session.modified = True
            logger.info(f'Members gate login successful: {username}')
            return jsonify({'success': True, 'message': 'Đăng nhập thành công'})
        logger.warning(f'Members gate login failed: username={username!r} (len_pwd={len(password)})')
        return (jsonify({'success': False, 'error': 'Tên đăng nhập hoặc mật khẩu không đúng. Vui lòng thử lại.'}), 401)
    except Exception as e:
        logger.error(f'Error in members_verify: {e}', exc_info=True)
        return (jsonify({'success': False, 'error': 'Lỗi server: ' + str(e)}), 500)


@members_portal_bp.route('/members/logout', methods=['GET', 'POST'])
def members_logout():
    """Đăng xuất khỏi cổng Members."""
    session.pop('members_gate_ok', None)
    session.pop('members_gate_user', None)
    logger.info('Members gate logout')
    from flask import redirect
    return redirect('/members')


@members_portal_bp.route('/api/members')
@rate_limit("120 per minute")
def get_members():
    """
    API lấy danh sách thành viên (source of truth).
    Yêu cầu session['members_gate_ok']. Có cache 5 phút.
    """
    from flask import current_app
    from mysql.connector import Error
    if not session.get('members_gate_ok'):
        logger.warning('Unauthorized access to /api/members')
        return (jsonify({'success': False, 'error': 'Chưa đăng nhập. Vui lòng đăng nhập lại.'}), 401)
    try:
        from extensions import cache
    except Exception:
        cache = None
    cache_key = 'api_members_data'
    if cache:
        try:
            cached_data = cache.get(cache_key)
            if cached_data is not None:
                return jsonify(cached_data)
        except Exception as e:
            logger.warning(f'Cache get error: {e}')
    from db import get_db_connection
    from services.person_service import load_relationship_data
    connection = None
    cursor = None
    try:
        connection = get_db_connection()
        if not connection:
            return (jsonify({'success': False, 'error': 'Không thể kết nối database'}), 500)
        cursor = connection.cursor(dictionary=True)
        cursor.execute("""
            SELECT COLUMN_NAME FROM information_schema.COLUMNS
            WHERE TABLE_SCHEMA = DATABASE() AND TABLE_NAME = 'persons'
            AND COLUMN_NAME IN ('csv_id', 'personal_image_url', 'personal_image', 'biography', 'academic_rank', 'academic_degree', 'phone', 'email', 'place_of_death', 'occupation')
        """)
        _col_rows = cursor.fetchall()
        available_columns = set()
        for row in _col_rows:
            col = row.get('COLUMN_NAME') or row.get('column_name') or ''
            if col:
                available_columns.add(str(col).strip())

        # Nhánh: ưu tiên persons.branch_name nếu có; nếu không thì join branches qua branch_id
        has_branch_name_col = False
        has_branch_id = False
        has_branches_table = False
        try:
            cursor.execute("""
                SELECT COLUMN_NAME FROM information_schema.COLUMNS
                WHERE TABLE_SCHEMA = DATABASE() AND TABLE_NAME = 'persons' AND COLUMN_NAME = 'branch_name'
                LIMIT 1
            """)
            has_branch_name_col = bool(cursor.fetchone())
            cursor.execute("""
                SELECT COLUMN_NAME FROM information_schema.COLUMNS
                WHERE TABLE_SCHEMA = DATABASE() AND TABLE_NAME = 'persons' AND COLUMN_NAME = 'branch_id'
                LIMIT 1
            """)
            has_branch_id = bool(cursor.fetchone())
            cursor.execute("""
                SELECT TABLE_NAME FROM information_schema.TABLES
                WHERE TABLE_SCHEMA = DATABASE() AND TABLE_NAME = 'branches'
                LIMIT 1
            """)
            has_branches_table = bool(cursor.fetchone())
        except Exception as e:
            logger.warning(f'Could not detect branch schema: {e}')
            has_branch_name_col = False
            has_branch_id = False
            has_branches_table = False

        select_fields = [
            'p.person_id', 'p.father_mother_id AS fm_id', 'p.full_name', 'p.alias', 'p.gender', 'p.status',
            'p.generation_level AS generation_number', 'p.birth_date_solar', 'p.birth_date_lunar',
            'p.death_date_solar', 'p.death_date_lunar', 'p.grave_info AS grave'
        ]
        select_fields.append('p.place_of_death' if 'place_of_death' in available_columns else 'NULL AS place_of_death')
        select_fields.append('p.personal_image_url AS personal_image_url' if 'personal_image_url' in available_columns else 'p.personal_image AS personal_image_url' if 'personal_image' in available_columns else 'NULL AS personal_image_url')
        select_fields.append('p.biography' if 'biography' in available_columns else 'NULL AS biography')
        select_fields.append('p.academic_rank' if 'academic_rank' in available_columns else 'NULL AS academic_rank')
        select_fields.append('p.academic_degree' if 'academic_degree' in available_columns else 'NULL AS academic_degree')
        select_fields.append('p.phone' if 'phone' in available_columns else 'NULL AS phone')
        select_fields.append('p.email' if 'email' in available_columns else 'NULL AS email')
        select_fields.append('p.occupation' if 'occupation' in available_columns else 'NULL AS occupation')
        select_fields.append('p.csv_id' if 'csv_id' in available_columns else 'NULL AS csv_id')
        if has_branch_name_col:
            select_fields.append('p.branch_name AS branch_name')
        else:
            select_fields.append('b.branch_name AS branch_name' if (has_branch_id and has_branches_table) else 'NULL AS branch_name')
        cursor.execute(f"""
            SELECT {', '.join(select_fields)}
            FROM persons p
            {'LEFT JOIN branches b ON p.branch_id = b.branch_id' if (has_branch_id and has_branches_table) else ''}
            ORDER BY COALESCE(p.generation_level, 999) ASC,
                CASE WHEN p.person_id LIKE 'P-%' AND SUBSTRING(p.person_id, 3) REGEXP '^[0-9]+-[0-9]+$'
                    THEN CAST(SUBSTRING_INDEX(SUBSTRING_INDEX(p.person_id, '-', 2), '-', -1) AS UNSIGNED) ELSE 999999 END ASC,
                CASE WHEN p.person_id LIKE 'P-%' AND SUBSTRING(p.person_id, 3) REGEXP '^[0-9]+-[0-9]+$'
                    THEN CAST(SUBSTRING_INDEX(p.person_id, '-', -1) AS UNSIGNED) ELSE 999999 END ASC,
                p.person_id ASC, p.full_name ASC
        """)
        persons = cursor.fetchall()
        relationship_data = load_relationship_data(cursor)
        spouse_data_from_table = relationship_data['spouse_data_from_table']
        spouse_data_from_marriages = relationship_data['spouse_data_from_marriages']
        spouse_data_from_csv = relationship_data['spouse_data_from_csv']
        parent_data = relationship_data['parent_data']
        children_map = relationship_data['children_map']
        siblings_map = relationship_data['siblings_map']
        members = []
        for person in persons:
            person_id = person['person_id']
            rel = parent_data.get(person_id, {'father_name': None, 'mother_name': None})
            spouse_names = spouse_data_from_table.get(person_id) or spouse_data_from_marriages.get(person_id) or spouse_data_from_csv.get(person_id) or []
            siblings = siblings_map.get(person_id, [])
            children = children_map.get(person_id, [])
            member = {
                'person_id': person_id, 'csv_id': person.get('csv_id') or person_id, 'fm_id': person.get('fm_id'), 'full_name': person.get('full_name'),
                'alias': person.get('alias'), 'gender': person.get('gender'), 'status': person.get('status'),
                'generation_number': person.get('generation_number'),
                'birth_date_solar': str(person['birth_date_solar']) if person.get('birth_date_solar') else None,
                'birth_date_lunar': str(person['birth_date_lunar']) if person.get('birth_date_lunar') else None,
                'death_date_solar': str(person['death_date_solar']) if person.get('death_date_solar') else None,
                'death_date_lunar': str(person['death_date_lunar']) if person.get('death_date_lunar') else None,
                'grave': person.get('grave'), 'grave_info': person.get('grave'), 'place_of_death': person.get('place_of_death'),
                'branch_name': person.get('branch_name'),
                'father_name': rel.get('father_name'), 'mother_name': rel.get('mother_name'),
                'spouses': '; '.join(spouse_names) if spouse_names else None,
                'siblings': '; '.join(siblings) if siblings else None,
                'children': '; '.join(children) if children else None,
                'personal_image_url': person.get('personal_image_url'), 'biography': person.get('biography'),
                'academic_rank': person.get('academic_rank'), 'academic_degree': person.get('academic_degree'),
                'phone': person.get('phone'), 'email': person.get('email'),
                'occupation': person.get('occupation')
            }
            members.append(member)
        response_data = {'success': True, 'data': members}
        if cache:
            try:
                cache.set(cache_key, response_data, timeout=300)
            except Exception as e:
                logger.warning(f'Cache set error: {e}')
        return jsonify(response_data)
    except Error as e:
        logger.error(f'Error in /api/members: {e}', exc_info=True)
        return (jsonify({'success': False, 'error': str(e)}), 500)
    except Exception as e:
        logger.error(f'Unexpected error in /api/members: {e}', exc_info=True)
        return (jsonify({'success': False, 'error': str(e)}), 500)
    finally:
        if cursor:
            try:
                cursor.close()
            except Exception:
                pass
        if connection and getattr(connection, 'is_connected', lambda: False)():
            try:
                connection.close()
            except Exception:
                pass


def _fetch_members_list():
    """
    Lấy danh sách thành viên (giống /api/members), không dùng cache.
    Returns (list of member dicts, None) hoặc (None, error_message).
    """
    from db import get_db_connection
    from services.person_service import load_relationship_data
    from mysql.connector import Error as MySqlError
    connection = None
    cursor = None
    try:
        connection = get_db_connection()
        if not connection:
            return (None, 'Không thể kết nối database')
        cursor = connection.cursor(dictionary=True)
        cursor.execute("""
            SELECT COLUMN_NAME FROM information_schema.COLUMNS
            WHERE TABLE_SCHEMA = DATABASE() AND TABLE_NAME = 'persons'
            AND COLUMN_NAME IN ('csv_id', 'personal_image_url', 'personal_image', 'biography', 'academic_rank', 'academic_degree', 'phone', 'email', 'place_of_death', 'occupation')
        """)
        _col_rows = cursor.fetchall()
        available_columns = set()
        for row in _col_rows:
            col = row.get('COLUMN_NAME') or row.get('column_name') or ''
            if col:
                available_columns.add(str(col).strip())

        # Nhánh: ưu tiên persons.branch_name nếu có; nếu không thì join branches qua branch_id
        has_branch_name_col = False
        has_branch_id = False
        has_branches_table = False
        try:
            cursor.execute("""
                SELECT COLUMN_NAME FROM information_schema.COLUMNS
                WHERE TABLE_SCHEMA = DATABASE() AND TABLE_NAME = 'persons' AND COLUMN_NAME = 'branch_name'
                LIMIT 1
            """)
            has_branch_name_col = bool(cursor.fetchone())
            cursor.execute("""
                SELECT COLUMN_NAME FROM information_schema.COLUMNS
                WHERE TABLE_SCHEMA = DATABASE() AND TABLE_NAME = 'persons' AND COLUMN_NAME = 'branch_id'
                LIMIT 1
            """)
            has_branch_id = bool(cursor.fetchone())
            cursor.execute("""
                SELECT TABLE_NAME FROM information_schema.TABLES
                WHERE TABLE_SCHEMA = DATABASE() AND TABLE_NAME = 'branches'
                LIMIT 1
            """)
            has_branches_table = bool(cursor.fetchone())
        except Exception as e:
            logger.warning(f'Could not detect branch schema (_fetch_members_list): {e}')
            has_branch_name_col = False
            has_branch_id = False
            has_branches_table = False

        select_fields = [
            'p.person_id', 'p.father_mother_id AS fm_id', 'p.full_name', 'p.alias', 'p.gender', 'p.status',
            'p.generation_level AS generation_number', 'p.birth_date_solar', 'p.birth_date_lunar',
            'p.death_date_solar', 'p.death_date_lunar', 'p.grave_info AS grave'
        ]
        select_fields.append('p.place_of_death' if 'place_of_death' in available_columns else 'NULL AS place_of_death')
        select_fields.append('p.personal_image_url AS personal_image_url' if 'personal_image_url' in available_columns else 'p.personal_image AS personal_image_url' if 'personal_image' in available_columns else 'NULL AS personal_image_url')
        select_fields.append('p.biography' if 'biography' in available_columns else 'NULL AS biography')
        select_fields.append('p.academic_rank' if 'academic_rank' in available_columns else 'NULL AS academic_rank')
        select_fields.append('p.academic_degree' if 'academic_degree' in available_columns else 'NULL AS academic_degree')
        select_fields.append('p.phone' if 'phone' in available_columns else 'NULL AS phone')
        select_fields.append('p.email' if 'email' in available_columns else 'NULL AS email')
        select_fields.append('p.occupation' if 'occupation' in available_columns else 'NULL AS occupation')
        select_fields.append('p.csv_id' if 'csv_id' in available_columns else 'NULL AS csv_id')
        if has_branch_name_col:
            select_fields.append('p.branch_name AS branch_name')
        else:
            select_fields.append('b.branch_name AS branch_name' if (has_branch_id and has_branches_table) else 'NULL AS branch_name')
        cursor.execute(f"""
            SELECT {', '.join(select_fields)}
            FROM persons p
            {'LEFT JOIN branches b ON p.branch_id = b.branch_id' if (has_branch_id and has_branches_table) else ''}
            ORDER BY COALESCE(p.generation_level, 999) ASC,
                CASE WHEN p.person_id LIKE 'P-%' AND SUBSTRING(p.person_id, 3) REGEXP '^[0-9]+-[0-9]+$'
                    THEN CAST(SUBSTRING_INDEX(SUBSTRING_INDEX(p.person_id, '-', 2), '-', -1) AS UNSIGNED) ELSE 999999 END ASC,
                CASE WHEN p.person_id LIKE 'P-%' AND SUBSTRING(p.person_id, 3) REGEXP '^[0-9]+-[0-9]+$'
                    THEN CAST(SUBSTRING_INDEX(p.person_id, '-', -1) AS UNSIGNED) ELSE 999999 END ASC,
                p.person_id ASC, p.full_name ASC
        """)
        persons = cursor.fetchall()
        relationship_data = load_relationship_data(cursor)
        parent_data = relationship_data['parent_data']
        spouse_data_from_table = relationship_data['spouse_data_from_table']
        spouse_data_from_marriages = relationship_data['spouse_data_from_marriages']
        spouse_data_from_csv = relationship_data['spouse_data_from_csv']
        children_map = relationship_data['children_map']
        siblings_map = relationship_data['siblings_map']
        members = []
        for person in persons:
            person_id = person['person_id']
            rel = parent_data.get(person_id, {'father_name': None, 'mother_name': None})
            spouse_names = spouse_data_from_table.get(person_id) or spouse_data_from_marriages.get(person_id) or spouse_data_from_csv.get(person_id) or []
            siblings = siblings_map.get(person_id, [])
            children = children_map.get(person_id, [])
            member = {
                'person_id': person_id, 'csv_id': person.get('csv_id') or person_id, 'fm_id': person.get('fm_id'), 'full_name': person.get('full_name'),
                'alias': person.get('alias'), 'gender': person.get('gender'), 'status': person.get('status'),
                'generation_number': person.get('generation_number'),
                'birth_date_solar': str(person['birth_date_solar']) if person.get('birth_date_solar') else None,
                'birth_date_lunar': str(person['birth_date_lunar']) if person.get('birth_date_lunar') else None,
                'death_date_solar': str(person['death_date_solar']) if person.get('death_date_solar') else None,
                'death_date_lunar': str(person['death_date_lunar']) if person.get('death_date_lunar') else None,
                'grave': person.get('grave'), 'grave_info': person.get('grave'), 'place_of_death': person.get('place_of_death'),
                'branch_name': person.get('branch_name'),
                'father_name': rel.get('father_name'), 'mother_name': rel.get('mother_name'),
                'spouses': '; '.join(spouse_names) if spouse_names else None,
                'siblings': '; '.join(siblings) if siblings else None,
                'children': '; '.join(children) if children else None,
                'personal_image_url': person.get('personal_image_url'), 'biography': person.get('biography'),
                'academic_rank': person.get('academic_rank'), 'academic_degree': person.get('academic_degree'),
                'phone': person.get('phone'), 'email': person.get('email'), 'occupation': person.get('occupation')
            }
            members.append(member)
        return (members, None)
    except MySqlError as e:
        logger.error(f'Error in _fetch_members_list: {e}', exc_info=True)
        return (None, str(e))
    except Exception as e:
        logger.error(f'Unexpected error in _fetch_members_list: {e}', exc_info=True)
        return (None, str(e))
    finally:
        if cursor:
            try:
                cursor.close()
            except Exception:
                pass
        if connection and getattr(connection, 'is_connected', lambda: False)():
            try:
                connection.close()
            except Exception:
                pass


# Cột xuất Excel: (key trong dict, tiêu đề tiếng Việt)
_EXCEL_COLUMNS = [
    ('person_id', 'ID'), ('fm_id', 'FM_ID'), ('branch_name', 'Nhánh'), ('full_name', 'Họ và tên'), ('alias', 'Tên gọi khác'),
    ('gender', 'Giới tính'), ('status', 'Trạng thái'), ('generation_number', 'Đời'),
    ('birth_date_solar', 'Ngày sinh (dương lịch)'), ('birth_date_lunar', 'Ngày sinh (âm lịch)'),
    ('death_date_solar', 'Ngày mất (dương lịch)'), ('death_date_lunar', 'Ngày mất (âm lịch)'),
    ('grave', 'Mộ'), ('place_of_death', 'Nơi mất'),
    ('father_name', 'Cha'), ('mother_name', 'Mẹ'),
    ('spouses', 'Vợ/Chồng'), ('siblings', 'Anh chị em'), ('children', 'Con'),
    ('occupation', 'Nghề nghiệp'), ('academic_rank', 'Học hàm'), ('academic_degree', 'Học vị'),
    ('phone', 'Điện thoại'), ('email', 'Email'), ('biography', 'Tiểu sử'), ('personal_image_url', 'URL ảnh'),
]


@members_portal_bp.route('/members/export/excel')
@rate_limit("20 per hour")
def export_members_excel():
    """Xuất toàn bộ danh sách thành viên ra file Excel. Yêu cầu đã đăng nhập cổng Members."""
    if not session.get('members_gate_ok'):
        from flask import redirect
        return redirect('/members')
    members, err = _fetch_members_list()
    if err:
        return (jsonify({'success': False, 'error': err}), 500)
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment
        from datetime import datetime
        wb = Workbook()
        ws = wb.active
        ws.title = 'Danh sách thành viên'
        # Header
        for col, (_, header) in enumerate(_EXCEL_COLUMNS, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center', wrap_text=True)
        # Data rows
        for row_idx, member in enumerate(members, 2):
            for col_idx, (key, _) in enumerate(_EXCEL_COLUMNS, 1):
                val = member.get(key)
                if val is not None:
                    ws.cell(row=row_idx, column=col_idx, value=str(val) if not isinstance(val, (int, float)) else val)
        # Auto-fit column widths (approximate)
        from openpyxl.utils import get_column_letter
        for col_idx in range(1, len(_EXCEL_COLUMNS) + 1):
            ws.column_dimensions[get_column_letter(col_idx)].width = 16
        buf = BytesIO()
        wb.save(buf)
        buf.seek(0)
        filename = 'danh_sach_thanh_vien_{}.xlsx'.format(datetime.now().strftime('%Y%m%d_%H%M'))
        return send_file(buf, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet', as_attachment=True, download_name=filename)
    except ImportError:
        logger.error('openpyxl not installed')
        return (jsonify({'success': False, 'error': 'Thư viện xuất Excel chưa được cài đặt.'}), 500)
    except Exception as e:
        logger.error(f'Error exporting Excel: {e}', exc_info=True)
        return (jsonify({'success': False, 'error': str(e)}), 500)


@members_portal_bp.route('/api/members/bulk-update-branch', methods=['POST'])
@rate_limit("30 per hour")
def bulk_update_members_branch():
    """
    Bulk update cột "Nhánh" cho members từ file Excel/CSV.

    Template file chỉ cần 2 cột:
    - ID: ứng với persons.person_id
    - Nhánh: chỉ nhận 1 trong {Một, Hai, Ba, Bốn}

    Rule:
    - Nếu Nhánh sai hoặc ID sai format => error_count++ và bỏ qua dòng
    - Nếu ID hợp lệ nhưng không tồn tại trong DB => silent (không update, không tính lỗi)
    - Trả về số dòng cập nhật thành công và số dòng lỗi.
    """
    import io
    import os
    import re
    import csv

    from db import get_db_connection
    from services.person_service import get_or_create_branch
    from services.members_service import get_members_password
    from utils.validation import secure_compare

    if not session.get('members_gate_ok'):
        logger.warning('Unauthorized access to /api/members/bulk-update-branch')
        return (jsonify({'success': False, 'error': 'Chưa đăng nhập. Vui lòng đăng nhập lại.'}), 401)

    # Auth: read password (multipart/form-data or JSON)
    if request.content_type and 'multipart/form-data' in request.content_type:
        password = (request.form.get('password') or '').strip()
    else:
        payload = request.get_json(silent=True) or {}
        password = (payload.get('password') or '').strip()

    correct_password = get_members_password()
    if not password or not secure_compare(password, correct_password):
        return (jsonify({'success': False, 'error': 'Mật khẩu không đúng hoặc chưa được cung cấp'}), 403)

    uploaded_file = request.files.get('file')
    if not uploaded_file or not uploaded_file.filename:
        return (jsonify({'success': False, 'error': 'Chưa nhận được file'}), 400)

    filename = uploaded_file.filename.strip()
    ext = os.path.splitext(filename)[1].lower()
    if ext not in {'.xlsx', '.csv'}:
        return (jsonify({'success': False, 'error': 'Chỉ hỗ trợ file .xlsx hoặc .csv'}), 400)

    # Cho phép cả dạng text (Một/Hai/...) và dạng code số (0..7, -1)
    branch_code_to_name = {
        '0': 'Tổ tiên',
        '1': 'Một',
        '2': 'Hai',
        '3': 'Ba',
        '4': 'Bốn',
        '5': 'Năm',
        '6': 'Sáu',
        '7': 'Bảy',
        '-1': 'Khác',
    }
    allowed_branch_names = set(branch_code_to_name.values())
    id_regex = re.compile(r'^P-\d+-\d+$')

    # Validate + map: {person_id: branch_name}
    valid_to_update = {}
    error_count = 0

    file_bytes = uploaded_file.read()
    if not file_bytes:
        return (jsonify({'success': False, 'error': 'File rỗng'}), 400)

    try:
        if ext == '.xlsx':
            from openpyxl import load_workbook

            buf = io.BytesIO(file_bytes)
            wb = load_workbook(filename=buf, read_only=True, data_only=True)
            ws = wb.active

            header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
            if not header_row:
                return (jsonify({'success': False, 'error': 'Không tìm thấy header trong file'}), 400)

            header_cells = [(str(c).strip() if c is not None else '') for c in header_row]
            id_idx = None
            branch_idx = None
            for i, h in enumerate(header_cells):
                if h == 'ID':
                    id_idx = i
                elif h == 'Nhánh':
                    branch_idx = i

            if id_idx is None or branch_idx is None:
                return (jsonify({'success': False, 'error': 'File phải có 2 cột: ID và Nhánh'}), 400)

            for row in ws.iter_rows(min_row=2, values_only=True):
                id_val = row[id_idx] if id_idx < len(row) else None
                branch_val = row[branch_idx] if branch_idx < len(row) else None

                id_str = str(id_val).strip() if id_val is not None else ''
                branch_str = str(branch_val).strip() if branch_val is not None else ''

                if not id_regex.match(id_str):
                    error_count += 1
                    continue
                canonical_branch = None
                if branch_str in allowed_branch_names:
                    canonical_branch = branch_str
                elif branch_str in branch_code_to_name:
                    canonical_branch = branch_code_to_name[branch_str]

                if not canonical_branch:
                    error_count += 1
                    continue

                valid_to_update[id_str] = canonical_branch

        else:
            # CSV
            text = file_bytes.decode('utf-8-sig', errors='replace')
            reader = csv.DictReader(io.StringIO(text))
            if not reader.fieldnames:
                return (jsonify({'success': False, 'error': 'Không tìm thấy header trong file CSV'}), 400)

            normalized = {(fn or '').strip(): fn for fn in reader.fieldnames if fn is not None}
            if 'ID' not in normalized or 'Nhánh' not in normalized:
                return (jsonify({'success': False, 'error': 'File phải có 2 cột: ID và Nhánh'}), 400)

            id_key = normalized['ID']
            branch_key = normalized['Nhánh']

            for row in reader:
                id_val = row.get(id_key)
                branch_val = row.get(branch_key)

                id_str = str(id_val).strip() if id_val is not None else ''
                branch_str = str(branch_val).strip() if branch_val is not None else ''

                if not id_regex.match(id_str):
                    error_count += 1
                    continue
                canonical_branch = None
                if branch_str in allowed_branch_names:
                    canonical_branch = branch_str
                elif branch_str in branch_code_to_name:
                    canonical_branch = branch_code_to_name[branch_str]

                if not canonical_branch:
                    error_count += 1
                    continue

                valid_to_update[id_str] = canonical_branch

    except Exception as e:
        logger.error(f'Failed to parse uploaded branch file: {e}', exc_info=True)
        return (jsonify({'success': False, 'error': f'Lỗi đọc file: {str(e)}'}), 400)

    connection = get_db_connection()
    if not connection:
        return (jsonify({'success': False, 'error': 'Không thể kết nối database'}), 500)

    cursor = None
    try:
        cursor = connection.cursor(dictionary=True)

        if not valid_to_update:
            return jsonify({'success': True, 'updated_count': 0, 'error_count': error_count})

        ids = list(valid_to_update.keys())
        placeholders = ','.join(['%s'] * len(ids))
        cursor.execute(f'SELECT person_id FROM persons WHERE person_id IN ({placeholders})', tuple(ids))
        existing_set = {r['person_id'] for r in cursor.fetchall() if r.get('person_id')}

        # Silent: only update if exists
        existing_map = {pid: valid_to_update[pid] for pid in ids if pid in existing_set}
        updated_count = len(existing_map)

        if updated_count == 0:
            return jsonify({'success': True, 'updated_count': 0, 'error_count': error_count})

        # Detect schema for branch_name / branch_id
        cursor.execute("""
            SELECT COLUMN_NAME
            FROM information_schema.COLUMNS
            WHERE TABLE_SCHEMA = DATABASE()
              AND TABLE_NAME = 'persons'
              AND COLUMN_NAME IN ('branch_name', 'branch_id')
        """)
        cols = {r.get('COLUMN_NAME') for r in cursor.fetchall() if r.get('COLUMN_NAME')}

        cursor.execute("""
            SELECT TABLE_NAME
            FROM information_schema.TABLES
            WHERE TABLE_SCHEMA = DATABASE()
              AND TABLE_NAME = 'branches'
            LIMIT 1
        """)
        has_branches_table = bool(cursor.fetchone())

        if 'branch_name' in cols:
            update_rows = [(branch, pid) for pid, branch in existing_map.items()]
            cursor.executemany(
                'UPDATE persons SET branch_name = %s WHERE person_id = %s',
                update_rows
            )
        elif 'branch_id' in cols and has_branches_table:
            # Map branch_name -> branch_id and update branch_id
            unique_branch_names = set(existing_map.values())
            branch_name_to_id = {}
            # get_or_create_branch() trong app.py kỳ vọng fetchone() trả về tuple
            # nên dùng cursor thường (không dictionary=True) để tránh KeyError.
            cursor_for_branch = connection.cursor()
            try:
                for bn in unique_branch_names:
                    branch_name_to_id[bn] = get_or_create_branch(cursor_for_branch, bn)
            finally:
                try:
                    cursor_for_branch.close()
                except Exception:
                    pass

            update_rows = [(branch_name_to_id[bn], pid) for pid, bn in existing_map.items()]
            cursor.executemany(
                'UPDATE persons SET branch_id = %s WHERE person_id = %s',
                update_rows
            )
        else:
            return (jsonify({'success': False, 'error': 'Không tìm thấy cột branch_name hoặc branch_id trong persons'}), 500)

        connection.commit()

        # Invalidate members cache
        try:
            from extensions import cache
            if cache:
                cache.delete('api_members_data')
        except Exception as e:
            logger.warning(f'Cache invalidation error (continuing): {e}')

        return jsonify({'success': True, 'updated_count': updated_count, 'error_count': error_count})

    except Exception as e:
        connection.rollback()
        logger.error(f'Error in bulk_update_members_branch: {e}', exc_info=True)
        return (jsonify({'success': False, 'error': str(e)}), 500)
    finally:
        if cursor:
            try:
                cursor.close()
            except Exception:
                pass
        if connection and getattr(connection, 'is_connected', lambda: False)():
            try:
                connection.close()
            except Exception:
                pass


def _sll_cell_nonempty(val):
    if val is None:
        return False
    if isinstance(val, str) and not val.strip():
        return False
    return True


def _sll_normalize_cell(val):
    if val is None:
        return None
    if isinstance(val, datetime):
        return val.strftime('%Y-%m-%d')
    if isinstance(val, date):
        return val.isoformat()
    if isinstance(val, float):
        if val == int(val):
            return int(val)
    return val


def _normalize_sll_row_id(val):
    """Chuẩn hóa ô cột ID từ Excel/CSV (float 1.0, khoảng trắng, v.v.)."""
    if val is None:
        return ''
    if isinstance(val, float):
        if val == int(val):
            return str(int(val))
        return str(val).strip()
    if isinstance(val, (datetime, date)):
        return str(val).strip()
    return str(val).strip()


def _sll_branch_code_to_name():
    return {
        '0': 'Tổ tiên',
        '1': 'Một',
        '2': 'Hai',
        '3': 'Ba',
        '4': 'Bốn',
        '5': 'Năm',
        '6': 'Sáu',
        '7': 'Bảy',
        '-1': 'Khác',
    }


def _sll_canonical_branch(val):
    if val is None:
        return None
    s = str(val).strip()
    if not s:
        return None
    m = _sll_branch_code_to_name()
    if s in m.values():
        return s
    if s in m:
        return m[s]
    return s


def _sll_base_payload(cursor, person_id, rel_data):
    cursor.execute('SELECT * FROM persons WHERE person_id = %s', (person_id,))
    p = cursor.fetchone()
    if not p:
        return None
    pid = person_id
    parent = rel_data['parent_data'].get(pid, {})
    spouse_names = (
        rel_data['spouse_data_from_table'].get(pid)
        or rel_data['spouse_data_from_marriages'].get(pid)
        or rel_data['spouse_data_from_csv'].get(pid)
        or []
    )
    children = rel_data['children_map'].get(pid, [])
    siblings = rel_data['siblings_map'].get(pid, [])

    def semi(xs):
        if not xs:
            return None
        if isinstance(xs, list):
            return '; '.join(xs)
        return str(xs)

    branch_name = p.get('branch_name')
    if not branch_name and p.get('branch_id'):
        cursor.execute('SELECT branch_name FROM branches WHERE branch_id = %s', (p['branch_id'],))
        br = cursor.fetchone()
        if br:
            branch_name = br.get('branch_name')

    def fmt_date(d):
        if d is None:
            return None
        if hasattr(d, 'isoformat'):
            s = d.isoformat()
            return s[:10] if len(s) >= 10 else s
        return str(d)

    fm = p.get('father_mother_id')
    if fm is None:
        fm = p.get('fm_id')

    return {
        'full_name': p.get('full_name'),
        'alias': p.get('alias'),
        'fm_id': fm,
        'gender': p.get('gender'),
        'status': p.get('status'),
        'generation_number': p.get('generation_level'),
        'branch_name': branch_name,
        'birth_date_solar': fmt_date(p.get('birth_date_solar')),
        'birth_date_lunar': fmt_date(p.get('birth_date_lunar')),
        'death_date_solar': fmt_date(p.get('death_date_solar')),
        'death_date_lunar': fmt_date(p.get('death_date_lunar')),
        'grave_info': p.get('grave_info'),
        'place_of_death': p.get('place_of_death'),
        'father_name': parent.get('father_name'),
        'mother_name': parent.get('mother_name'),
        'spouse_info': semi(spouse_names),
        'children_info': semi(children),
        'siblings_info': semi(siblings),
        'occupation': p.get('occupation'),
        'academic_rank': p.get('academic_rank'),
        'academic_degree': p.get('academic_degree'),
        'phone': p.get('phone'),
        'email': p.get('email'),
        'biography': p.get('biography'),
        'personal_image_url': p.get('personal_image_url') or p.get('personal_image'),
    }


def _sll_merge_excel_into_payload(base, excel_by_internal_key):
    """Chỉ ghi đè các ô Excel khác rỗng; map spouses/children/siblings/grave -> đúng key form."""
    out = dict(base)
    key_map = {
        'spouses': 'spouse_info',
        'children': 'children_info',
        'siblings': 'siblings_info',
        'grave': 'grave_info',
    }
    for k, v in excel_by_internal_key.items():
        if k == 'person_id':
            continue
        if not _sll_cell_nonempty(v):
            continue
        nv = _sll_normalize_cell(v)
        pk = key_map.get(k, k)
        if pk == 'branch_name':
            nv = _sll_canonical_branch(nv)
        out[pk] = nv
    return out


def _normalize_excel_header(header):
    if header is None:
        return ''
    s = str(header).strip().lower()
    if not s:
        return ''
    # Bỏ dấu tiếng Việt để map được các biến thể header.
    s = ''.join(ch for ch in unicodedata.normalize('NFD', s) if unicodedata.category(ch) != 'Mn')
    s = s.replace('đ', 'd')
    # Gom mọi ký tự ngăn cách về 1 dấu cách.
    s = re.sub(r'[^a-z0-9]+', ' ', s)
    s = re.sub(r'\s+', ' ', s).strip()
    return s


_EXCEL_HEADER_ALIASES = {
    'person_id': ['ID'],
    'fm_id': ['FM_ID', 'FM-ID', 'FM ID', 'Father_Mother_ID', 'Father-Mother-ID', 'Father Mother ID'],
    'branch_name': ['Nhánh', 'Nhanh'],
    'full_name': ['Họ và tên', 'Ho va ten', 'Họ tên', 'Ho ten'],
    'alias': ['Tên gọi khác', 'Ten goi khac', 'Tên khác', 'Ten khac'],
    'gender': ['Giới tính', 'Gioi tinh'],
    'status': ['Trạng thái', 'Trang thai'],
    'generation_number': ['Đời', 'Doi'],
    'birth_date_solar': ['Ngày sinh (dương lịch)', 'Ngay sinh duong lich', 'Ngay sinh (duong lich)'],
    'birth_date_lunar': ['Ngày sinh (âm lịch)', 'Ngay sinh am lich', 'Ngay sinh (am lich)'],
    'death_date_solar': ['Ngày mất (dương lịch)', 'Ngay mat duong lich', 'Ngay mat (duong lich)'],
    'death_date_lunar': ['Ngày mất (âm lịch)', 'Ngay mat am lich', 'Ngay mat (am lich)'],
    'grave': ['Mộ', 'Mo'],
    'place_of_death': ['Nơi mất', 'Noi mat'],
    'father_name': ['Cha', 'Tên bố', 'Ten bo', 'Bo'],
    'mother_name': ['Mẹ', 'Tên mẹ', 'Ten me', 'Me'],
    'spouses': ['Vợ/Chồng', 'Vo/Chong', 'Vo chong', 'Hôn phối', 'Hon phoi'],
    # File thực tế đôi khi ghi nhầm "Con cái anh em".
    'siblings': ['Anh chị em', 'Anh chi em', 'Con cái anh em', 'Con cai anh em'],
    'children': ['Con', 'Con cái', 'Con cai'],
    'occupation': ['Nghề nghiệp', 'Nghe nghiep'],
    'academic_rank': ['Học hàm', 'Hoc ham'],
    'academic_degree': ['Học vị', 'Hoc vi'],
    'phone': ['Điện thoại', 'Dien thoai'],
    'email': ['Email'],
    'biography': ['Tiểu sử', 'Tieu su'],
    'personal_image_url': ['URL ảnh', 'Url anh', 'URL anh', 'Personal image URL'],
}


_EXCEL_HEADER_TO_KEY = {}
for _key, _aliases in _EXCEL_HEADER_ALIASES.items():
    for _alias in _aliases:
        _EXCEL_HEADER_TO_KEY[_normalize_excel_header(_alias)] = _key
# Giữ tương thích ngược với cặp key/header chuẩn.
for _key, _header in _EXCEL_COLUMNS:
    _EXCEL_HEADER_TO_KEY[_normalize_excel_header(_header)] = _key


@members_portal_bp.route('/members/template/Template_updatetbqc.xlsx')
def download_template_update_sll():
    """File mẫu Update SLL (cùng cấu trúc Xuất Excel)."""
    if not session.get('members_gate_ok'):
        return redirect('/members')
    root = Path(__file__).resolve().parent.parent
    path = root / 'Template_updatetbqc.xlsx'
    if not path.is_file():
        return (jsonify({'success': False, 'error': 'Không tìm thấy Template_updatetbqc.xlsx trên server'}), 404)
    return send_file(path, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet', as_attachment=True, download_name='Template_updatetbqc.xlsx')


@members_portal_bp.route('/api/members/bulk-update-sll', methods=['POST'])
@rate_limit("15 per hour")
def bulk_update_members_sll():
    """
    Cập nhật nhiều thành viên từ Excel/CSV (cùng template Xuất Excel / Template_updatetbqc.xlsx).
    Mỗi dòng: ID bắt buộc; ô trống = giữ nguyên giá trị hiện có trong DB (merge).
    """
    import csv
    import io

    from db import get_db_connection
    from services.members_service import get_members_password
    from services.person_service import apply_person_members_update_core, load_relationship_data
    from utils.validation import secure_compare

    if not session.get('members_gate_ok'):
        logger.warning('Unauthorized access to /api/members/bulk-update-sll')
        return (jsonify({'success': False, 'error': 'Chưa đăng nhập. Vui lòng đăng nhập lại.'}), 401)

    if not BULK_UPDATE_SLL_ENABLED:
        return (
            jsonify(
                {
                    'success': False,
                    'error': 'Chức năng Update SLL đang tạm dừng. Vui lòng dùng Thêm hoặc Cập nhật từng thành viên.',
                }
            ),
            503,
        )

    if request.content_type and 'multipart/form-data' in request.content_type:
        password = (request.form.get('password') or '').strip()
    else:
        payload = request.get_json(silent=True) or {}
        password = (payload.get('password') or '').strip()

    correct_password = get_members_password()
    if not password or not secure_compare(password, correct_password):
        return (jsonify({'success': False, 'error': 'Mật khẩu không đúng hoặc chưa được cung cấp'}), 403)

    uploaded_file = request.files.get('file')
    if not uploaded_file or not uploaded_file.filename:
        return (jsonify({'success': False, 'error': 'Chưa nhận được file'}), 400)

    filename = uploaded_file.filename.strip()
    ext = os.path.splitext(filename)[1].lower()
    if ext not in {'.xlsx', '.csv'}:
        return (jsonify({'success': False, 'error': 'Chỉ hỗ trợ file .xlsx hoặc .csv'}), 400)

    file_bytes = uploaded_file.read()
    if not file_bytes:
        return (jsonify({'success': False, 'error': 'File rỗng'}), 400)

    id_regex = re.compile(r'^P-\d+-\d+$')
    rows_to_process = []

    try:
        if ext == '.xlsx':
            from openpyxl import load_workbook

            buf = io.BytesIO(file_bytes)
            wb = load_workbook(filename=buf, read_only=True, data_only=True)
            ws = wb.active
            header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
            if not header_row:
                return (jsonify({'success': False, 'error': 'Không tìm thấy header trong file'}), 400)
            header_cells = [(str(c).strip() if c is not None else '') for c in header_row]
            idx_map = {}
            for i, h in enumerate(header_cells):
                hk = _normalize_excel_header(h)
                if hk in _EXCEL_HEADER_TO_KEY:
                    idx_map[_EXCEL_HEADER_TO_KEY[hk]] = i
            if 'person_id' not in idx_map:
                return (jsonify({'success': False, 'error': 'File phải có cột ID (đúng template Xuất Excel)'}), 400)
            for row in ws.iter_rows(min_row=2, values_only=True):
                row = list(row) if row else []
                id_val = row[idx_map['person_id']] if idx_map['person_id'] < len(row) else None
                id_str = _normalize_sll_row_id(id_val)
                if not id_str:
                    continue
                excel_dict = {}
                for key, col_idx in idx_map.items():
                    cell = row[col_idx] if col_idx < len(row) else None
                    excel_dict[key] = cell
                rows_to_process.append((id_str, excel_dict))
        else:
            text = file_bytes.decode('utf-8-sig', errors='replace')
            reader = csv.DictReader(io.StringIO(text))
            if not reader.fieldnames:
                return (jsonify({'success': False, 'error': 'Không tìm thấy header trong file CSV'}), 400)
            norm = {_normalize_excel_header(fn): fn for fn in reader.fieldnames if fn is not None}
            person_id_header = norm.get(_normalize_excel_header('ID'))
            if not person_id_header:
                return (jsonify({'success': False, 'error': 'File phải có cột ID'}), 400)
            id_key = person_id_header
            for r in reader:
                id_val = r.get(id_key)
                id_str = _normalize_sll_row_id(id_val)
                if not id_str:
                    continue
                excel_dict = {}
                for header_norm, ikey in _EXCEL_HEADER_TO_KEY.items():
                    if header_norm not in norm:
                        continue
                    excel_dict[ikey] = r.get(norm[header_norm])
                rows_to_process.append((id_str, excel_dict))

    except Exception as e:
        logger.error(f'Failed to parse SLL file: {e}', exc_info=True)
        return (jsonify({'success': False, 'error': f'Lỗi đọc file: {str(e)}'}), 400)

    connection = get_db_connection()
    if not connection:
        return (jsonify({'success': False, 'error': 'Không thể kết nối database'}), 500)

    cursor = None
    updated_count = 0
    error_count = 0
    skipped_count = 0

    try:
        cursor = connection.cursor(dictionary=True)
        rel_data = load_relationship_data(cursor)
        has_csv_id_col = False
        try:
            cursor.execute(
                """
                SELECT COLUMN_NAME
                FROM information_schema.COLUMNS
                WHERE TABLE_SCHEMA = DATABASE()
                  AND TABLE_NAME = 'persons'
                  AND COLUMN_NAME = 'csv_id'
                LIMIT 1
                """
            )
            has_csv_id_col = bool(cursor.fetchone())
        except Exception as schema_err:
            logger.warning(f'bulk-update-sll: could not detect csv_id column: {schema_err}')
            has_csv_id_col = False

        for id_str, excel_dict in rows_to_process:
            target_person_id = None

            if id_regex.match(id_str):
                cursor.execute('SELECT person_id FROM persons WHERE person_id = %s', (id_str,))
                found = cursor.fetchone()
                if found:
                    target_person_id = found.get('person_id')

            # Fallback theo csv_id để hỗ trợ file có ID hiển thị khác person_id.
            if not target_person_id and has_csv_id_col:
                cursor.execute('SELECT person_id FROM persons WHERE csv_id = %s LIMIT 1', (id_str,))
                found = cursor.fetchone()
                if found:
                    target_person_id = found.get('person_id')

            if not target_person_id:
                if id_regex.match(id_str) or has_csv_id_col:
                    skipped_count += 1
                else:
                    error_count += 1
                continue

            if not id_regex.match(target_person_id):
                skipped_count += 1
                continue

            base = _sll_base_payload(cursor, target_person_id, rel_data)
            if not base:
                skipped_count += 1
                continue
            merged = _sll_merge_excel_into_payload(base, excel_dict)
            try:
                ok, err, _code = apply_person_members_update_core(connection, cursor, target_person_id, merged, None, None)
                if ok:
                    updated_count += 1
                    rel_data = load_relationship_data(cursor)
                else:
                    error_count += 1
                    logger.warning(f'bulk-update-sll row {id_str} ({target_person_id}): {err}')
                    try:
                        connection.rollback()
                    except Exception:
                        pass
            except Exception as row_err:
                try:
                    connection.rollback()
                except Exception:
                    pass
                error_count += 1
                logger.warning(f'bulk-update-sll row {id_str} ({target_person_id}) exception: {row_err}', exc_info=True)

        try:
            from extensions import cache
            if cache:
                cache.delete('api_members_data')
        except Exception as e:
            logger.warning(f'Cache invalidation error (bulk SLL): {e}')

        return jsonify({
            'success': True,
            'updated_count': updated_count,
            'error_count': error_count,
            'skipped_count': skipped_count,
        })

    except Exception as e:
        logger.error(f'Error in bulk_update_members_sll: {e}', exc_info=True)
        return (jsonify({'success': False, 'error': str(e)}), 500)
    finally:
        if cursor:
            try:
                cursor.close()
            except Exception:
                pass
        if connection and getattr(connection, 'is_connected', lambda: False)():
            try:
                connection.close()
            except Exception:
                pass
