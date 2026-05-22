# -*- coding: utf-8 -*-
"""
Phase 5.4 — Members export contract tests.

DB integration tests cho GET /members/export/excel.
Không thay đổi production code. Chỉ DB state assertion + response contract.
"""
import pytest
from io import BytesIO


# ── helpers ───────────────────────────────────────────────────────────────────


def _commit(cursor):
    conn = getattr(cursor, "_connection", None)
    if conn is None:
        raise AssertionError("cursor does not expose _connection")
    conn.commit()


def _seed_person(cursor, person_id, full_name, generation_level=None, status=None):
    cursor.execute(
        "INSERT INTO persons (person_id, full_name, generation_level, status) VALUES (%s, %s, %s, %s)",
        (person_id, full_name, generation_level, status),
    )
    _commit(cursor)


def _set_members_gate(db_client):
    with db_client.session_transaction() as sess:
        sess["members_gate_ok"] = True
        sess["members_gate_user"] = "pytest"


def _parse_xlsx(data: bytes):
    """Load openpyxl workbook từ raw bytes."""
    from openpyxl import load_workbook
    return load_workbook(BytesIO(data))


XLSX_MIME = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"


# ── auth gate ─────────────────────────────────────────────────────────────────


@pytest.mark.db_integration
def test_export_no_gate_redirects(db_client, test_db_cursor):
    """Không có session gate → redirect về /members (302)."""
    resp = db_client.get("/members/export/excel")
    assert resp.status_code == 302
    location = resp.headers.get("Location", "")
    assert "/members" in location


# ── empty DB ──────────────────────────────────────────────────────────────────


@pytest.mark.db_integration
def test_export_empty_db_returns_valid_xlsx(db_client, test_db_cursor):
    """DB không có persons → vẫn trả 200 và file xlsx hợp lệ."""
    _set_members_gate(db_client)
    resp = db_client.get("/members/export/excel")
    assert resp.status_code == 200
    assert XLSX_MIME in resp.content_type
    wb = _parse_xlsx(resp.data)
    assert wb.active is not None


@pytest.mark.db_integration
def test_export_empty_db_has_header_only(db_client, test_db_cursor):
    """DB không có persons → sheet chỉ có 1 hàng (header), không có data row."""
    _set_members_gate(db_client)
    resp = db_client.get("/members/export/excel")
    assert resp.status_code == 200
    ws = _parse_xlsx(resp.data).active
    assert ws.max_row == 1


@pytest.mark.db_integration
def test_export_header_contains_expected_columns(db_client, test_db_cursor):
    """Header row phải chứa các cột cốt lõi: ID, Họ và tên, Giới tính."""
    _set_members_gate(db_client)
    resp = db_client.get("/members/export/excel")
    assert resp.status_code == 200
    ws = _parse_xlsx(resp.data).active
    headers = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
    for expected in ("ID", "Họ và tên", "Giới tính"):
        assert expected in headers, f"Header '{expected}' không có trong xlsx"


# ── data rows ─────────────────────────────────────────────────────────────────


@pytest.mark.db_integration
def test_export_row_count_matches_seeded_persons(db_client, test_db_cursor):
    """Seed 3 persons → xlsx có 4 hàng (1 header + 3 data)."""
    _seed_person(test_db_cursor, "P-TEST-1", "Nguyễn Văn A", generation_level=1)
    _seed_person(test_db_cursor, "P-TEST-2", "Trần Thị B", generation_level=2)
    _seed_person(test_db_cursor, "P-TEST-3", "Lê Văn C", generation_level=2)

    _set_members_gate(db_client)
    resp = db_client.get("/members/export/excel")
    assert resp.status_code == 200
    ws = _parse_xlsx(resp.data).active
    assert ws.max_row == 4  # 1 header + 3 data


@pytest.mark.db_integration
def test_export_person_name_present_in_xlsx(db_client, test_db_cursor):
    """Tên người được seed phải xuất hiện đúng trong xlsx."""
    _seed_person(test_db_cursor, "P-EXPORT-9", "Đinh Thị Kiểm Tra")

    _set_members_gate(db_client)
    resp = db_client.get("/members/export/excel")
    assert resp.status_code == 200
    ws = _parse_xlsx(resp.data).active

    # Collect all cell values in the sheet
    all_values = {ws.cell(row=r, column=c).value for r in range(1, ws.max_row + 1) for c in range(1, ws.max_column + 1)}
    assert "Đinh Thị Kiểm Tra" in all_values


@pytest.mark.db_integration
def test_export_content_disposition_is_attachment(db_client, test_db_cursor):
    """Response phải có Content-Disposition: attachment với tên file xlsx."""
    _set_members_gate(db_client)
    resp = db_client.get("/members/export/excel")
    assert resp.status_code == 200
    cd = resp.headers.get("Content-Disposition", "")
    assert "attachment" in cd
    assert ".xlsx" in cd
